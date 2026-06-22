"""Field-trial CSV loading, weather XLSX conversion, train/test split."""
from __future__ import annotations

from pathlib import Path

import numpy as np
import openpyxl
import pandas as pd
from openpyxl import Workbook


BIO_COLUMNS = ["LeavesDW", "StemDW", "tubersDW", "LAI"]
ROOT_COLUMN = "rootsDW"

PCSE_COLS = ["DAY", "IRRAD", "TMIN", "TMAX", "VAP", "WIND", "RAIN", "SNOWDEPTH"]
UNITS = ["date", "kJ/m2/day or hours", "Celsius", "Celsius", "kPa", "m/sec", "mm", "cm"]


def convert_weather_to_pcse_format(src, dst, *, force=False):
    src = Path(src); dst = Path(dst)
    if dst.exists() and not force:
        return dst
    src_wb = openpyxl.load_workbook(src, data_only=True)
    sh = src_wb.active
    country, station, description = sh.cell(2, 2).value, sh.cell(3, 2).value, sh.cell(4, 2).value
    source_text, nodata_val = sh.cell(5, 2).value, sh.cell(6, 2).value
    longitude, latitude, elevation = sh.cell(8, 1).value, sh.cell(8, 2).value, sh.cell(8, 3).value
    angstrom_a, angstrom_b, has_sunshine = sh.cell(8, 4).value, sh.cell(8, 5).value, sh.cell(8, 6).value
    df = pd.read_excel(src, header=9).iloc[1:].reset_index(drop=True)[PCSE_COLS].copy()
    df["DAY"] = pd.to_datetime(df["DAY"])
    for c in PCSE_COLS[1:]:
        df[c] = pd.to_numeric(df[c], errors="coerce")
    df.loc[df["SNOWDEPTH"] < 0, "SNOWDEPTH"] = nodata_val

    wb = Workbook(); s = wb.active; s.title = "Weather"
    s.cell(1, 1).value = "Site Characteristics"
    s.cell(2, 1).value = "Country"; s.cell(2, 2).value = country
    s.cell(3, 1).value = "Station"; s.cell(3, 2).value = station
    s.cell(4, 1).value = "Description"; s.cell(4, 2).value = description
    s.cell(5, 1).value = "Source"; s.cell(5, 2).value = source_text
    s.cell(6, 1).value = "Contact"
    s.cell(7, 1).value = "Missing values"; s.cell(7, 2).value = nodata_val
    for j, l in enumerate(["Longitude", "Latitude", "Elevation", "AngstromA", "AngstromB", "HasSunshine"], 1):
        s.cell(8, j).value = l
    for j, v in enumerate([longitude, latitude, elevation, angstrom_a, angstrom_b, has_sunshine], 1):
        s.cell(9, j).value = v
    s.cell(10, 1).value = "Observed data"
    for j, c in enumerate(PCSE_COLS, 1):
        s.cell(11, j).value = c
    for j, u in enumerate(UNITS, 1):
        s.cell(12, j).value = u
    for i, row in enumerate(df.itertuples(index=False), 13):
        for j, v in enumerate(row, 1):
            s.cell(i, j).value = v.to_pydatetime() if hasattr(v, "to_pydatetime") else v
    wb.save(dst)
    return dst


def convert_weather_files(weather_xlsx_by_site, data_temp_dir):
    """Convert raw weather XLSX files to PCSE-compatible format. Returns {site: pcse_path}."""
    data_temp_dir = Path(data_temp_dir)
    out = {}
    for site, src in weather_xlsx_by_site.items():
        dst = data_temp_dir / f"Weatherfile_{site}_pcse.xlsx"
        out[site] = convert_weather_to_pcse_format(src, dst)
    return out


def load_observations(csv_path):
    """Load field-trial observations and return (obs_df, PLOT_KEYS).

    Drops the small Shadow side-experiment, keeps only rows with at least one
    biomass/LAI observation, and builds the per-plot key list.
    """
    obs_df = pd.read_csv(csv_path)
    obs_df["Date"] = pd.to_datetime(obs_df["Date"])
    obs_df = obs_df[obs_df["Shadow"].isna()].copy()
    obs_df = obs_df[obs_df[BIO_COLUMNS + [ROOT_COLUMN]].notna().any(axis=1)].copy()

    plot_keys = (
        obs_df[["Year", "Location", "Plotnumber", "Cultivar", "Nitrogen", "Irrigation"]]
        .drop_duplicates()
        .sort_values(["Year", "Location", "Plotnumber"])
        .reset_index(drop=True)
    )
    PLOT_KEYS = list(plot_keys.itertuples(index=False, name="Plot"))
    return obs_df, PLOT_KEYS, plot_keys


def split_random(plots, test_fraction=0.2, seed=42):
    rng = np.random.default_rng(seed)
    idx = rng.permutation(len(plots))
    n_test = int(round(len(plots) * test_fraction))
    test_idx = set(idx[:n_test].tolist())
    train, test = [], []
    for i, p in enumerate(plots):
        (test if i in test_idx else train).append(p)
    return train, test
